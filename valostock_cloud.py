import streamlit as st
import pandas as pd
import os
from datetime import datetime
import matplotlib.pyplot as plt
from pathlib import Path
import smtplib
from email.message import EmailMessage
from openpyxl import load_workbook

st.set_page_config(page_title="Valorisation des Stocks", layout="wide")
st.title("📊 Dashboard - Valorisation des stocks (version Cloud)")

HISTO_FILE = Path("historique_valorisation.csv")
TCD_MODEL = Path("modele_valorisation_TCD.xlsx")
EXCEL_TCD_FILE = Path("rapport_valo_dynamique.xlsx")

if HISTO_FILE.exists():
    historique_df = pd.read_csv(HISTO_FILE)
else:
    historique_df = pd.DataFrame(columns=["date", "organisationId", "brand", "valorisation"])

st.sidebar.header("📂 Importer les fichiers")
stock_files = st.sidebar.file_uploader("Fichiers d'export Keyneo (un par magasin)", type=["csv"], accept_multiple_files=True)
product_file = st.sidebar.file_uploader("Base produit (Excel)", type=["xls", "xlsx"])

view_mode = st.sidebar.radio("Vue du rapport", ["Par fournisseur puis magasin", "Par magasin puis fournisseur"])

if stock_files and product_file:
    stock_list = [pd.read_csv(file, sep=';') for file in stock_files]
    stocks_df = pd.concat(stock_list, ignore_index=True)
    products_df = pd.read_excel(product_file)
    products_df = products_df.rename(columns={"SKU": "sku", "PurchasingPrice": "purchasing_price", "Brand": "brand"})
    products_df = products_df[["sku", "purchasing_price", "brand"]]
    stocks_df["sku"] = stocks_df["sku"].astype(str)
    products_df["sku"] = products_df["sku"].astype(str)

    merged_df = pd.merge(stocks_df, products_df, on="sku", how="left")
    merged_df["valorisation"] = merged_df["quantity"] * merged_df["purchasing_price"]
    date_import = datetime.today().strftime('%Y-%m-%d')

    report_df = merged_df.groupby(["organisationId", "brand"], as_index=False)["valorisation"].sum()
    report_df.insert(0, "date", date_import)
    report_df = report_df[report_df["valorisation"] > 0].drop_duplicates()
    report_df["valorisation"] = report_df["valorisation"].round(2)

    historique_df = historique_df[~(
        (historique_df["date"] == date_import) &
        (historique_df["organisationId"].isin(report_df["organisationId"])) &
        (historique_df["brand"].isin(report_df["brand"]))
    )]
    historique_df = pd.concat([historique_df, report_df], ignore_index=True)
    historique_df.to_csv(HISTO_FILE, index=False)

    
    if TCD_MODEL.exists():
        from shutil import copyfile
        import openpyxl
        from openpyxl.utils import get_column_letter

        copyfile(TCD_MODEL, EXCEL_TCD_FILE)
        wb = openpyxl.load_workbook(EXCEL_TCD_FILE)
        ws = wb["Données"]

        # Réécriture à partir de la ligne 2
        for i, row in enumerate(historique_df.itertuples(index=False), start=2):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j).value = value

        # Supprimer les lignes en trop
        if ws.max_row > len(historique_df) + 1:
            ws.delete_rows(len(historique_df) + 2, ws.max_row - len(historique_df) - 1)

        # Mettre à jour dynamiquement la plage de la table
        if "TableauStocks" in ws.tables:
            table = ws.tables["TableauStocks"]
            start_cell = "A1"
            end_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
            table.ref = f"{start_cell}:{end_cell}"

        wb.save(EXCEL_TCD_FILE)


    st.sidebar.header("🔍 Filtres")
    dates = historique_df["date"].unique().tolist()
    magasins = historique_df["organisationId"].unique().tolist()
    fournisseurs = historique_df["brand"].unique().tolist()
    selected_date = st.sidebar.selectbox("Date", sorted(dates, reverse=True))
    selected_magasin = st.sidebar.multiselect("Magasins", magasins, default=magasins)
    selected_brand = st.sidebar.multiselect("Fournisseurs", fournisseurs, default=fournisseurs)

    filtered_df = historique_df[
        (historique_df["date"] == selected_date) &
        (historique_df["organisationId"].isin(selected_magasin)) &
        (historique_df["brand"].isin(selected_brand))
    ]

    if view_mode == "Par fournisseur puis magasin":
        filtered_df = filtered_df.sort_values(by=["brand", "organisationId", "valorisation"], ascending=[True, True, False])
    else:
        filtered_df = filtered_df.sort_values(by=["organisationId", "brand", "valorisation"], ascending=[True, True, False])

    # 📈 Graphique dynamique dans l'app
    st.subheader("📈 Évolution de la valorisation par magasin pour un fournisseur")
    filtered_brands = historique_df.groupby("brand")["valorisation"].sum()
    filtered_brands = filtered_brands[filtered_brands > 0].index.tolist()

    selected_graph_brand = st.selectbox("Choisir un fournisseur", sorted(filtered_brands))
    selected_graph_magasin = st.multiselect(
        "Choisir un ou plusieurs magasins",
        sorted(historique_df["organisationId"].dropna().unique()),
        default=sorted(historique_df["organisationId"].dropna().unique())
    )

    graph_data = historique_df[
        (historique_df["brand"] == selected_graph_brand) &
        (historique_df["organisationId"].isin(selected_graph_magasin))
    ]

    if not graph_data.empty:
        fig, ax = plt.subplots()
        for magasin in graph_data["organisationId"].unique():
            data = graph_data[graph_data["organisationId"] == magasin]
            grouped = data.groupby("date")["valorisation"].sum().reset_index()
            ax.plot(grouped["date"], grouped["valorisation"], marker='o', label=magasin)

        ax.set_title(f"Évolution du stock de {selected_graph_brand} par magasin")
        ax.set_xlabel("Date")
        ax.set_ylabel("Valorisation (€)")
        ax.legend()
        ax.tick_params(axis='x', rotation=45)
        st.pyplot(fig)

    # 📬 Envoi par mail avec fichier dynamique joint
    st.subheader("✉️ Envoi par e-mail")
    emails_supp = st.text_input("Autres destinataires (séparés par des virgules)")
    if st.button("📧 M'envoyer ce rapport par e-mail"):
        smtp_server = st.secrets["email"]["smtp_server"]
        smtp_port = st.secrets["email"]["smtp_port"]
        smtp_user = st.secrets["email"]["smtp_user"]
        smtp_password = st.secrets["email"]["smtp_password"]
        receiver = st.secrets["email"]["receiver"]

        all_recipients = [receiver] + [email.strip() for email in emails_supp.split(",") if email.strip() != ""]

        msg = EmailMessage()
        msg["Subject"] = f"Rapport valorisation stocks - {selected_date}"
        msg["From"] = smtp_user
        msg["To"] = ", ".join(all_recipients)
        msg.set_content(
    f"""Bonjour,

Veuillez trouver ci-joint le rapport dynamique de valorisation des stocks à jour pour la date {selected_date}.
"""
)

        if EXCEL_TCD_FILE.exists():
            with open(EXCEL_TCD_FILE, "rb") as f:
                msg.add_attachment(f.read(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=EXCEL_TCD_FILE.name)

        try:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
                st.success("📨 Rapport dynamique envoyé avec succès par e-mail !")
        except Exception as e:
            st.error(f"Erreur lors de l'envoi de l'e-mail : {e}")

    # 🗓️ Historique
    st.subheader("🗓️ Historique complet")
    st.dataframe(historique_df.sort_values("date", ascending=False), use_container_width=True)

else:
    st.info("Merci d'importer les fichiers de stock et la base produit pour générer le rapport.")
